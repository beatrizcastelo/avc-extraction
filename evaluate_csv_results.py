#!/usr/bin/env python3
"""
evaluate_csv_results.py
Lê CSVs de extracção por modelo de csv_results/, aplica limpeza regex
e compara métricas (Prec, Rec, F1, MAE) antes vs. depois, por modelo.

Colunas obrigatórias no CSV:
    case_id
    model
    extracted_fields.extracted key     — nome do campo (ex: nihss_admissao)
    extracted_fields.extracted value   — output raw do LLM (ex: "10 pontos")
    ground_truth                       — valor correcto esperado
    exact_match                        — 0/1 antes de limpeza
    fuzzy_match                        — 0/1 antes de limpeza

Uso:
    python evaluate_csv_results.py
    python evaluate_csv_results.py --input csv_results/
    python evaluate_csv_results.py --output csv_eval_reports/
    python evaluate_csv_results.py --no-report
"""

import re
import sys
import argparse
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd

# ── Classificação de campos ────────────────────────────────────────────────────

TIMESTAMP_VARS   = {"sintomas","admissaoorigem","admissaocoimbra","tcce",
                    "fibrinolise","transferencia","puncaofemoral","recanalizacao"}
METRIC_VARS      = {"onset_to_door","door_to_imaging","door_to_needle",
                    "door_to_puncture","onset_to_recan","door_in_door_out","door1_to_door2"}
SCALE_VARS       = {"nihss_admissao","nihss_alta","mrs_previo","mrs_alta","mrs_3meses",
                    "nihss_admissao_carta","nihss_alta_carta","mrs_previo_carta",
                    "mrs_alta_carta","mrs_3meses_carta","mrs_3meses_consulta"}
CATEGORICAL_VARS = {"tipo","etiologia_toast","tratamento","territorio","complicacoes","causa_obito"}
BINARY_VARS      = {"vivo_30_dias"}
NUMERIC_VARS     = {"dias_obito"}

NULL_TOKENS = {"n/a","na","null","none","não aplicável","desconhecido",""}

TIMESTAMP_TOL_MIN = 1   # tolerância exact_match para timestamps
METRIC_TOL_MIN    = 5   # tolerância exact_match para métricas temporais
SCALE_FUZZY_TOL   = 1   # fuzzy_match para escalas (±1 ponto)
TIMESTAMP_FUZZY   = 5   # fuzzy_match para timestamps (±5 min)
METRIC_FUZZY      = 15  # fuzzy_match para métricas (±15 min)
CAT_FUZZY_RATIO   = 0.8 # SequenceMatcher threshold para categóricas


# ── Limpeza regex ──────────────────────────────────────────────────────────────

def clean_numeric(val):
    """Extrai o primeiro número da string. 'NIHSS: 10 pontos' → 10."""
    if val is None or str(val).strip().lower() in NULL_TOKENS:
        return None
    try:
        n = float(val)
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        m = re.search(r"(\d+(?:\.\d+)?)", str(val))
        if m:
            n = float(m.group(1))
            return int(n) if n == int(n) else n
        return None


def clean_timestamp(val):
    """'10h30', '10H30', '10:30' → '10:30'. Deixa formatos ISO intactos."""
    if val is None or str(val).strip().lower() in NULL_TOKENS:
        return None
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})[hH:](\d{2})$', s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def clean_metric(val):
    """'1h30', '90', '1:30' → 90 (minutos inteiros)."""
    if val is None or str(val).strip().lower() in NULL_TOKENS:
        return None
    s = str(val).strip()
    m = re.match(r'^(\d+)[hH:](\d{2})$', s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def clean_categorical(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    for a, b in [("á","a"),("à","a"),("â","a"),("ã","a"),("é","e"),("ê","e"),
                 ("í","i"),("ó","o"),("ô","o"),("õ","o"),("ú","u"),("ç","c")]:
        s = s.replace(a, b)
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'[_\-]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def clean_binary(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in {"true","sim","yes","1","vivo"}:
        return True
    if s in {"false","não","nao","no","0","faleceu","óbito","obito"}:
        return False
    return None


def get_field_type(field: str) -> str:
    f = str(field).lower().strip()
    if f in SCALE_VARS:      return "scale"
    if f in TIMESTAMP_VARS:  return "timestamp"
    if f in METRIC_VARS:     return "metric"
    if f in CATEGORICAL_VARS: return "categorical"
    if f in BINARY_VARS:     return "binary"
    if f in NUMERIC_VARS:    return "numeric"
    return "unknown"


def apply_cleaning(val, field_type: str):
    if field_type in ("scale", "numeric"):
        return clean_numeric(val)
    if field_type == "timestamp":
        return clean_timestamp(val)
    if field_type == "metric":
        return clean_metric(val)
    if field_type == "categorical":
        return clean_categorical(val)
    if field_type == "binary":
        return clean_binary(val)
    # unknown: tenta numérico primeiro, cai para string
    n = clean_numeric(val)
    return n if n is not None else (str(val).strip() if val is not None else None)


# ── Comparadores pós-limpeza ───────────────────────────────────────────────────

def is_null(v) -> bool:
    return v is None or str(v).strip().lower() in NULL_TOKENS


def _parse_ts(s) -> "datetime | None":
    for fmt in ("%H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M",
                "%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except ValueError:
            continue
    return None


def compare_after_cleaning(pred_clean, gt_raw, field_type: str) -> dict:
    """Calcula exact_match e fuzzy_match depois de limpar pred e gt."""
    gt_clean = apply_cleaning(gt_raw, field_type)

    pn, gn = is_null(pred_clean), is_null(gt_clean)
    if gn and pn:     return {"em": 0, "fm": 0, "tn": 1, "mae": None}
    if gn and not pn: return {"em": 0, "fm": 0, "tn": 0, "mae": None}
    if not gn and pn: return {"em": 0, "fm": 0, "tn": 0, "mae": None}

    if field_type in ("scale", "numeric"):
        try:
            vp, vg = float(pred_clean), float(gt_clean)
            mae = abs(vp - vg)
            return {"em": int(mae == 0), "fm": int(mae <= SCALE_FUZZY_TOL), "tn": 0, "mae": mae}
        except (ValueError, TypeError):
            return {"em": 0, "fm": 0, "tn": 0, "mae": None}

    if field_type == "timestamp":
        tp, tg = _parse_ts(pred_clean), _parse_ts(gt_clean)
        if tp is None or tg is None:
            return {"em": 0, "fm": 0, "tn": 0, "mae": None}
        diff = abs((tp - tg).total_seconds()) / 60
        return {"em": int(diff <= TIMESTAMP_TOL_MIN),
                "fm": int(diff <= TIMESTAMP_FUZZY),
                "tn": 0, "mae": diff}

    if field_type == "metric":
        try:
            vp, vg = int(pred_clean), int(clean_metric(gt_raw))
            mae = abs(vp - vg)
            return {"em": int(mae <= METRIC_TOL_MIN),
                    "fm": int(mae <= METRIC_FUZZY),
                    "tn": 0, "mae": float(mae)}
        except (ValueError, TypeError):
            return {"em": 0, "fm": 0, "tn": 0, "mae": None}

    if field_type == "binary":
        if pred_clean is None or gt_clean is None:
            return {"em": 0, "fm": 0, "tn": 0, "mae": None}
        hit = (pred_clean == gt_clean)
        return {"em": int(hit), "fm": int(hit), "tn": 0, "mae": None}

    # categorical / unknown
    ps = str(pred_clean) if pred_clean is not None else ""
    gs = str(gt_clean)   if gt_clean   is not None else ""
    exact = (ps == gs)
    ratio = SequenceMatcher(None, ps, gs).ratio()
    return {"em": int(exact), "fm": int(ratio >= CAT_FUZZY_RATIO), "tn": 0, "mae": None}


# ── Cálculo de métricas ────────────────────────────────────────────────────────

def aggregate_metrics(rows: list[dict],
                      match_col: str,
                      val_col: str,
                      mae_col: str | None = None) -> dict:
    """
    Agrega TP/FP/FN e calcula Prec/Rec/F1/MAE.
    Distingue TN (ambos null) de erros.
    """
    tp = fp = fn = tn = 0
    mae_vals = []

    for r in rows:
        gt_null = is_null(r.get("ground_truth"))
        pr_null = is_null(r.get(val_col))
        em      = bool(int(r.get(match_col, 0) or 0))

        if gt_null and pr_null:
            tn += 1
        elif gt_null and not pr_null:
            fp += 1
        elif not gt_null and pr_null:
            fn += 1
        elif em:
            tp += 1
        else:
            fp += 1
            fn += 1

        if mae_col and r.get(mae_col) is not None:
            try:
                mae_vals.append(float(r[mae_col]))
            except (ValueError, TypeError):
                pass

    prec   = tp / (tp + fp) if (tp + fp) > 0 else None
    recall = tp / (tp + fn) if (tp + fn) > 0 else None
    f1     = (2 * prec * recall / (prec + recall)
              if prec is not None and recall is not None and (prec + recall) > 0
              else None)
    mae    = sum(mae_vals) / len(mae_vals) if mae_vals else None

    return {
        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
        "precision": round(prec,   4) if prec   is not None else None,
        "recall":    round(recall, 4) if recall is not None else None,
        "F1":        round(f1,     4) if f1     is not None else None,
        "MAE":       round(mae,    2) if mae    is not None else None,
        "n":         tp + fp + fn + tn,
    }


# ── Pipeline ───────────────────────────────────────────────────────────────────

REQUIRED_COLS = {
    "case_id",
    "model",
    "extracted_fields.extracted key",
    "extracted_fields.extracted value",
    "ground_truth",
    "exact_match",
    "fuzzy_match",
}

KEY_COL = "extracted_fields.extracted key"
VAL_COL = "extracted_fields.extracted value"


def load_csvs(folder: Path) -> pd.DataFrame:
    csvs = sorted(folder.glob("*.csv"))
    if not csvs:
        print(f"❌  Nenhum CSV encontrado em: {folder}")
        sys.exit(1)

    frames = []
    for f in csvs:
        try:
            df = pd.read_csv(f)
        except Exception as e:
            print(f"  ⚠️  {f.name} ignorado — erro de leitura: {e}")
            continue
        missing = REQUIRED_COLS - set(df.columns)
        if missing:
            print(f"  ⚠️  {f.name} ignorado — colunas em falta: {missing}")
            continue
        df["_source_file"] = f.name
        frames.append(df)
        print(f"  ✅  {f.name}  ({len(df)} linhas)")

    if not frames:
        print("❌  Nenhum CSV válido encontrado.")
        sys.exit(1)

    return pd.concat(frames, ignore_index=True)


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    """Adiciona colunas: tipo de campo, valor limpo, métricas pós-limpeza."""
    df = df.copy()

    df["_field_type"]    = df[KEY_COL].apply(lambda x: get_field_type(str(x)))
    df["_cleaned_value"] = df.apply(
        lambda r: apply_cleaning(r[VAL_COL], r["_field_type"]), axis=1
    )

    results = df.apply(
        lambda r: compare_after_cleaning(r["_cleaned_value"], r["ground_truth"], r["_field_type"]),
        axis=1,
    )
    df["exact_match_clean"] = [r["em"]  for r in results]
    df["fuzzy_match_clean"] = [r["fm"]  for r in results]
    df["_mae_clean"]        = [r["mae"] for r in results]
    df["_tn_clean"]         = [r["tn"]  for r in results]

    return df


def build_detail(df: pd.DataFrame) -> pd.DataFrame:
    """Tabela detalhada por (model, field) com métricas before e after."""
    rows = []
    for (model, field, ftype), grp in df.groupby(["model", KEY_COL, "_field_type"]):
        grp_rows = grp.to_dict("records")
        before = aggregate_metrics(grp_rows, "exact_match",       VAL_COL)
        after  = aggregate_metrics(grp_rows, "exact_match_clean", "_cleaned_value",
                                   mae_col="_mae_clean")
        row = {"model": model, "field": field, "group": ftype}
        for k, v in before.items():
            row[f"{k}_before"] = v
        for k, v in after.items():
            row[f"{k}_after"] = v
        for metric in ("F1", "precision", "recall", "MAE"):
            b = before.get(metric)
            a = after.get(metric)
            if b is not None and a is not None:
                row[f"Δ{metric}"] = round(a - b, 4)
            else:
                row[f"Δ{metric}"] = None
        rows.append(row)

    return pd.DataFrame(rows).sort_values(["model", "field"])


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Resumo por (model, group) — útil para ver qual modelo beneficia mais."""
    rows = []
    for (model, group), grp in df.groupby(["model", "_field_type"]):
        grp_rows = grp.to_dict("records")
        before = aggregate_metrics(grp_rows, "exact_match",       VAL_COL)
        after  = aggregate_metrics(grp_rows, "exact_match_clean", "_cleaned_value",
                                   mae_col="_mae_clean")
        rows.append({
            "model":       model,
            "group":       group,
            "n":           before["n"],
            "F1_before":   before["F1"],
            "F1_after":    after["F1"],
            "ΔF1":         round((after["F1"] or 0) - (before["F1"] or 0), 4),
            "Prec_before": before["precision"],
            "Prec_after":  after["precision"],
            "Rec_before":  before["recall"],
            "Rec_after":   after["recall"],
            "MAE_before":  before["MAE"],
            "MAE_after":   after["MAE"],
        })

    return pd.DataFrame(rows).sort_values(["model", "group"])


# ── Output ─────────────────────────────────────────────────────────────────────

def print_report(df_summary: pd.DataFrame):
    sep = "=" * 85
    print(f"\n{sep}")
    print("  IMPACTO DA LIMPEZA REGEX POR MODELO")
    print(sep)

    for model, grp in df_summary.groupby("model"):
        print(f"\n  Modelo: {model}")
        header = f"  {'Grupo':<15} {'F1_antes':>9} {'F1_depois':>9} {'ΔF1':>8}  "
        header += f"{'MAE_antes':>10} {'MAE_depois':>10}  {'n':>5}"
        print(header)
        print("  " + "-" * 73)
        for _, row in grp.iterrows():
            f1b = f"{row.F1_before:9.4f}" if row.F1_before is not None else "      n/a"
            f1a = f"{row.F1_after:9.4f}"  if row.F1_after  is not None else "      n/a"
            df1 = f"{row.ΔF1:+8.4f}"      if pd.notna(row.get("ΔF1")) else "     n/a"
            mb  = f"{row.MAE_before:10.2f}" if row.MAE_before is not None else "       n/a"
            ma  = f"{row.MAE_after:10.2f}"  if row.MAE_after  is not None else "       n/a"
            print(f"  {row.group:<15} {f1b} {f1a} {df1}  {mb} {ma}  {int(row.n):>5}")

    # Top melhorias
    top = df_summary[df_summary["ΔF1"].notna() & (df_summary["ΔF1"] > 0)].sort_values("ΔF1", ascending=False).head(10)
    if not top.empty:
        print(f"\n  TOP melhorias com limpeza regex (ΔF1 > 0):")
        print(f"  {'Modelo':<30} {'Grupo':<15} {'ΔF1':>8}")
        print("  " + "-" * 55)
        for _, row in top.iterrows():
            print(f"  {row.model:<30} {row.group:<15} {row.ΔF1:>+8.4f}")

    # Modelos sem melhoria (podem estar ok, ou ter problemas diferentes)
    no_gain = df_summary[df_summary["ΔF1"].notna() & (df_summary["ΔF1"] <= 0)]
    if not no_gain.empty:
        print(f"\n  Grupos sem ganho (ΔF1 ≤ 0):")
        print(f"  {'Modelo':<30} {'Grupo':<15} {'ΔF1':>8} {'F1_depois':>9}")
        print("  " + "-" * 65)
        for _, row in no_gain.sort_values("ΔF1").iterrows():
            f1a = f"{row.F1_after:9.4f}" if row.F1_after is not None else "      n/a"
            print(f"  {row.model:<30} {row.group:<15} {row.ΔF1:>+8.4f} {f1a}")

    print(f"\n{sep}\n")


def save_reports(df_enriched: pd.DataFrame,
                 df_detail: pd.DataFrame,
                 df_summary: pd.DataFrame,
                 output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    paths = {
        "detail":  output_dir / f"eval_detail_{ts}.csv",
        "summary": output_dir / f"eval_summary_{ts}.csv",
        "rows":    output_dir / f"eval_rows_{ts}.csv",
        "excel":   output_dir / f"eval_report_{ts}.xlsx",
    }

    df_detail.to_csv(paths["detail"], index=False)
    df_summary.to_csv(paths["summary"], index=False)

    export_cols = [
        "case_id", "model", KEY_COL, VAL_COL, "_cleaned_value",
        "ground_truth", "exact_match", "exact_match_clean",
        "fuzzy_match", "fuzzy_match_clean", "_mae_clean", "_field_type",
    ]
    df_enriched[[c for c in export_cols if c in df_enriched.columns]].to_csv(
        paths["rows"], index=False
    )

    try:
        with pd.ExcelWriter(paths["excel"], engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Resumo por Modelo", index=False)
            df_detail.to_excel(writer, sheet_name="Detalhe por Campo", index=False)
            df_enriched[[c for c in export_cols if c in df_enriched.columns]].to_excel(
                writer, sheet_name="Linhas", index=False
            )

            from openpyxl.styles import PatternFill, Font
            green  = PatternFill("solid", fgColor="C6EFCE")
            yellow = PatternFill("solid", fgColor="FFEB9C")
            red    = PatternFill("solid", fgColor="FFC7CE")
            blue   = PatternFill("solid", fgColor="DDEBF7")

            for sheet_name, df_sheet, metric_cols in [
                ("Resumo por Modelo", df_summary, ["F1_before","F1_after"]),
                ("Detalhe por Campo", df_detail,  ["F1_before","F1_after"]),
            ]:
                ws = writer.sheets[sheet_name]
                cols = list(df_sheet.columns)
                for col_name in metric_cols:
                    if col_name not in cols:
                        continue
                    col_idx = cols.index(col_name) + 1
                    for row_idx in range(2, len(df_sheet) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val  = cell.value
                        if isinstance(val, (int, float)):
                            cell.fill = (green if val >= 0.9
                                         else yellow if val >= 0.7
                                         else red)

                # Destaca ΔF1 positivo a azul
                for col_name in [c for c in cols if c.startswith("Δ")]:
                    col_idx = cols.index(col_name) + 1
                    for row_idx in range(2, len(df_sheet) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.fill = blue
                            cell.font = Font(bold=True)

        print(f"  Excel:   {paths['excel']}")
    except Exception as e:
        print(f"  ⚠️  Excel não gerado: {e}")

    print(f"  Detail:  {paths['detail']}")
    print(f"  Summary: {paths['summary']}")
    print(f"  Rows:    {paths['rows']}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Avalia CSVs de extracção por modelo após limpeza regex"
    )
    parser.add_argument("--input",     default="csv_results",
                        help="Pasta com os CSVs de entrada (default: csv_results/)")
    parser.add_argument("--output",    default="csv_eval_reports",
                        help="Pasta para os relatórios (default: csv_eval_reports/)")
    parser.add_argument("--no-report", action="store_true",
                        help="Não guarda ficheiros de output")
    args = parser.parse_args()

    base    = Path(__file__).parent
    in_dir  = base / args.input
    out_dir = base / args.output

    print(f"\n📂 A ler CSVs de: {in_dir}")
    df = load_csvs(in_dir)
    print(f"\n   {len(df)} linhas | "
          f"{df['model'].nunique()} modelos | "
          f"{df[KEY_COL].nunique()} campos únicos\n")

    print("⚙️  A aplicar limpeza regex...")
    df_enriched = enrich(df)

    changed = (df_enriched["exact_match"].astype(int) != df_enriched["exact_match_clean"]).sum()
    print(f"   {changed} linhas com exact_match alterado após limpeza "
          f"({changed/len(df_enriched)*100:.1f}%)\n")

    df_detail  = build_detail(df_enriched)
    df_summary = build_summary(df_enriched)

    print_report(df_summary)

    if not args.no_report:
        print("📄 A guardar relatórios...")
        save_reports(df_enriched, df_detail, df_summary, out_dir)
        print()


if __name__ == "__main__":
    main()
