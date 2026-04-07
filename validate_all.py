#!/usr/bin/env python3
"""
validate_all.py — Validação Automática dos 150 Casos
Tese de Mestrado IA & Ciência de Dados — FCTUC / ULS Coimbra
Orientador: Prof. Pedro Furtado

Uso:
    python validate_all.py --cases 5
    python validate_all.py --backend groq
    python validate_all.py --backend ollama
    python validate_all.py --use-cache
"""

import json, os, re, sys, time, argparse
from unittest import result
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Any

sys.path.insert(0, "streamlit/agents")
import metrics


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

BASE_DIR      = Path(__file__).parent
DATA_DIR      = Path("/Users/beatrizcastelo/Documents/GitHub/groq-teste-processor/outputs_teste")
STREAMLIT_DIR = BASE_DIR / "streamlit"
OUTPUTS_DIR   = STREAMLIT_DIR / "outputs"
REPORT_DIR    = BASE_DIR / "validation_reports"
REPORT_DIR.mkdir(exist_ok=True)

TIMESTAMP_TOLERANCE_MIN = 1
METRIC_TOLERANCE_MIN    = 5

NULL_TOKENS = {"n/a","na","null","none","não aplicável","desconhecido",""}

TIMESTAMP_VARS   = ["sintomas","admissaoorigem","admissaocoimbra","tcce",
                    "fibrinolise","transferencia","puncaofemoral","recanalizacao"]
METRIC_VARS      = ["onset_to_door","door_to_imaging","door_to_needle",
                    "door_to_puncture","onset_to_recan","door_in_door_out","door1_to_door2"]
SCALE_VARS       = ["nihss_admissao_carta","nihss_alta_carta","mrs_previo_carta",
                    "mrs_alta_carta","mrs_3meses_carta",
                    "nihss_admissao_consulta","nihss_alta_consulta",
                    "mrs_previo_consulta","mrs_alta_consulta","mrs_3meses_consulta"]
CATEGORICAL_VARS = ["tipo","etiologia_toast","tratamento","territorio","complicacoes","causa_obito"]
BINARY_VARS      = ["vivo_30_dias"]
NUMERIC_VARS     = ["dias_obito"]

ALL_VARS = (TIMESTAMP_VARS + METRIC_VARS + SCALE_VARS +
            CATEGORICAL_VARS + BINARY_VARS + NUMERIC_VARS)


# ══════════════════════════════════════════════════════════════════════════════
# FLATTEN
# ══════════════════════════════════════════════════════════════════════════════

def _extract_scale_value(section: dict, key: str):
    entry = section.get(key)
    if entry is None:
        return None
    v = entry.get("value") if isinstance(entry, dict) else entry
    if v is None or str(v).lower() in {"null","none","n/a","na",""}:
        return None
    try:
        num = float(v)
        return int(num) if num == int(num) else num
    except (ValueError, TypeError):
        return None


def flatten_extractor_output(raw: dict) -> dict:
    result = {}
    ts = raw.get("timestamps", {})

    def get_val(key):
        entry = ts.get(key, {})
        v = entry.get("value") if isinstance(entry, dict) else entry
        if v is None or str(v).strip().upper() in {"NA","N/A","NULL","NONE",""}:
            return None
        return str(v).strip()

    # Timestamps
    result["sintomas"]        = get_val("onset_uvb")
    result["admissaoorigem"]  = get_val("door1_admission")
    result["admissaocoimbra"] = get_val("admission")
    result["tcce"]            = get_val("imaging_ct")
    result["fibrinolise"]     = get_val("thrombolysis")
    result["transferencia"]   = get_val("door1_departure")
    result["puncaofemoral"]   = get_val("femoral_puncture")
    result["recanalizacao"]   = get_val("recanalization")
    door2 = get_val("door2")
    if door2 is not None:
        result["admissaocoimbra"] = door2

    # Métricas
    for k, v in raw.get("metricas_temporais", {}).items():
        if k in METRIC_VARS:
            result[k] = v
    for k, v in raw.get("metrics", {}).items():
        if isinstance(v, dict) and v.get("value") is not None:
            result[k] = v["value"]

    # Escalas
    scales = raw.get("scales", {})
    for source in ["carta", "consulta"]:
        s     = scales.get(source, {})
        nihss = s.get("nihss", {})
        mrs   = s.get("mrs", {})
        result[f"nihss_admissao_{source}"] = _extract_scale_value(nihss, "nihss_admissao")
        result[f"nihss_alta_{source}"]     = _extract_scale_value(nihss, "nihss_alta")
        result[f"mrs_previo_{source}"]     = _extract_scale_value(mrs,   "mrs_previo")
        result[f"mrs_alta_{source}"]       = _extract_scale_value(mrs,   "mrs_alta")
        result[f"mrs_3meses_{source}"]     = _extract_scale_value(mrs,   "mrs_3meses")

    # Categóricas (RF6)
    cat = raw.get("categorical", {})
    for var in ["tipo","etiologia_toast","tratamento","territorio","complicacoes"]:
        v = cat.get(var)
        result[var] = v if v not in (None, "", "null", "none") else None

    # Mortalidade
    mort = raw.get("mortality", {})
    vivo_raw = mort.get("vivo_30_dias")
    if isinstance(vivo_raw, bool):
        result["vivo_30_dias"] = vivo_raw
    elif isinstance(vivo_raw, str):
        result["vivo_30_dias"] = (True  if vivo_raw.lower() in {"true","sim","yes"} else
                                  False if vivo_raw.lower() in {"false","nao","não","no"} else None)
    else:
        result["vivo_30_dias"] = None

    result["dias_obito"]  = mort.get("dias_obito")
    result["causa_obito"] = mort.get("causa_obito")

    return result


def flatten_ground_truth(raw: dict) -> dict:
    result = {}

    ts = raw.get("timestamps", {})
    for gt_key, var in [("sintomas","sintomas"),("admissao_coimbra","admissaocoimbra"),
                        ("admissao_origem","admissaoorigem"),("tc_ce","tcce"),
                        ("fibrinolise","fibrinolise"),("transferencia","transferencia"),
                        ("puncao_femoral","puncaofemoral"),("recanalizacao","recanalizacao")]:
        v = ts.get(gt_key)
        result[var] = str(v).strip() if v is not None else None

    mt = raw.get("metricas_temporais", {})
    for gt_key, var in [("onset_to_door","onset_to_door"),("door_to_imaging","door_to_imaging"),
                        ("door_to_needle","door_to_needle"),("door_to_puncture","door_to_puncture"),
                        ("onset_to_recan","onset_to_recan"),("door_in_door_out","door_in_door_out"),
                        ("door1_to_door2","door1_to_door2")]:
        v = mt.get(gt_key)
        result[var] = str(v).strip() if v is not None else None

    result["nihss_admissao_carta"] = raw.get("nihss_admissao")
    result["nihss_alta_carta"]     = raw.get("nihss_alta")
    result["mrs_previo_carta"]     = raw.get("mrs_previo")
    result["mrs_alta_carta"]       = raw.get("mrs_alta")

    seg = raw.get("seguimento", {})
    result["vivo_30_dias"]     = seg.get("vivo_30_dias")
    result["dias_obito"]       = seg.get("dias_obito")
    result["causa_obito"]      = seg.get("causa_obito")
    result["mrs_3meses_carta"] = seg.get("mrs_3_meses")

    for var in ["tipo","etiologia_toast","tratamento","territorio","complicacoes"]:
        result[var] = raw.get(var)

    consulta_gt = raw.get("consulta", {})
    result["nihss_admissao_consulta"] = consulta_gt.get("nihss_admissao")
    result["nihss_alta_consulta"]     = consulta_gt.get("nihss_alta")
    result["mrs_previo_consulta"]     = consulta_gt.get("mrs_previo")
    result["mrs_alta_consulta"]       = consulta_gt.get("mrs_alta")
    result["mrs_3meses_consulta"]     = consulta_gt.get("mrs_3_meses")

    return result


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE PARSING
# ══════════════════════════════════════════════════════════════════════════════

def is_null(value: Any) -> bool:
    return value is None or str(value).strip().lower() in NULL_TOKENS

def normalize_cat(value: Any) -> str:
    if value is None: return ""
    s = str(value).strip().lower()
    # remove acentos comuns para comparação robusta
    for a, b in [("á","a"),("à","a"),("â","a"),("ã","a"),("é","e"),("ê","e"),
                 ("í","i"),("ó","o"),("ô","o"),("õ","o"),("ú","u"),("ç","c")]:
        s = s.replace(a, b)
    return s.replace("-"," ").replace("_"," ")

# Mapeamento de variantes de etiologia_toast para forma canónica normalizada
_ETIOLOGIA_ALIASES = {
    "cardioembólica":                     "cardioembólica",
    "cardioembólico":                     "cardioembólica",
    "cardioembolica":                     "cardioembólica",
    "cardioembólico":                     "cardioembólica",
    "aterosclerose grandes vasos":        "aterosclerose grandes vasos",
    "aterotrombótica de grande vaso":     "aterosclerose grandes vasos",
    "aterotrombotica de grande vaso":     "aterosclerose grandes vasos",
    "aterotrombótico de grande vaso":     "aterosclerose grandes vasos",
    "aterosclerose de grandes vasos":     "aterosclerose grandes vasos",
    "oclusão pequenos vasos (lacunar)":   "oclusao pequenos vasos",
    "oclusao pequenos vasos (lacunar)":   "oclusao pequenos vasos",
    "oclusão de pequeno vaso":            "oclusao pequenos vasos",
    "oclusao de pequeno vaso":            "oclusao pequenos vasos",
    "pequeno vaso":                       "oclusao pequenos vasos",
    "lacunar":                            "oclusao pequenos vasos",
    "indeterminado":                      "indeterminado",
    "etiologia indeterminada":            "indeterminado",
    "indeterminada":                      "indeterminado",
    "outra etiologia determinada":        "outra etiologia determinada",
    "outras etiologias determinadas":     "outra etiologia determinada",
}

def normalize_etiologia(value: Any) -> str:
    if value is None: return ""
    s = normalize_cat(value)
    return _ETIOLOGIA_ALIASES.get(s, s)

def _tratamento_key(value: Any) -> str:
    """
    Extrai a 'chave semântica' do tratamento para comparação fuzzy.
    Casos:
      - fibrinólise pré-hospitalar simples  → "fibrinolise"
      - bridging (fibrinólise + TEV)        → "bridging:[cidade]"
      - TEV isolada contraindicação         → "tev_contraindicacao"
      - TEV isolada fora de janela          → "tev_fora_janela"
      - conservador                         → "conservador"
    """
    if value is None: return ""
    s = normalize_cat(str(value))

    if "conservador" in s or "sem reperfusao" in s:
        return "conservador"
    if "fora de janela" in s or "wake" in s:
        return "tev_fora_janela"
    if "contraindicac" in s:
        return "tev_contraindicacao"
    if "trombectomia" in s and ("+" in s or "em " in s):
        # bridging: extrai cidade de origem se possível
        m = re.search(r"fibrinolise em ([a-z ]+?) \+", s)
        cidade = m.group(1).strip() if m else "?"
        return f"bridging:{cidade}"
    if "fibrinolise" in s or "rt-pa" in s or "tenecteplase" in s or "alteplase" in s:
        return "fibrinolise"
    if "tev" in s or "trombectomia" in s:
        return "tev"
    return s

def parse_hhhmm(value: Any) -> int | None:
    if is_null(value): return None
    s = str(value).strip()
    m = re.match(r"^(\d+)[hH:](\d{2})$", s)
    if m: return int(m.group(1)) * 60 + int(m.group(2))
    if re.match(r"^\d+$", s): return int(s)
    return None

def parse_timestamp(value: Any) -> datetime | None:
    if is_null(value): return None
    s = str(value).strip()
    m = re.match(r"^(\d{1,2})[hH](\d{2})$", s)
    if m:
        try: return datetime.strptime(f"{m.group(1)}:{m.group(2)}", "%H:%M")
        except ValueError: pass
    for fmt in ("%Y-%m-%dT%H:%M","%Y-%m-%d %H:%M","%d/%m/%Y %H:%M","%d-%m-%Y %H:%M","%H:%M"):
        try: return datetime.strptime(s, fmt)
        except ValueError: continue
    return None


# ══════════════════════════════════════════════════════════════════════════════
# COMPARADORES
# ══════════════════════════════════════════════════════════════════════════════

def compare_timestamp(pred, gt) -> dict:
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1,"mae_min":None}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0,"mae_min":None}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0,"mae_min":None}
    tp_dt, gt_dt = parse_timestamp(str(pred)), parse_timestamp(str(gt))
    if tp_dt is None or gt_dt is None:
        return {"tp":0,"fp":1,"fn":1,"tn":0,"mae_min":None}
    diff = abs((tp_dt - gt_dt).total_seconds()) / 60
    hit  = diff <= TIMESTAMP_TOLERANCE_MIN
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0,"mae_min":diff}

def compare_metric(pred, gt) -> dict:
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1,"mae_min":None}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0,"mae_min":None}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0,"mae_min":None}
    vp, vg = parse_hhhmm(str(pred)), parse_hhhmm(str(gt))
    if vp is None or vg is None:
        return {"tp":0,"fp":1,"fn":1,"tn":0,"mae_min":None}
    mae = abs(vp - vg)
    hit = mae <= METRIC_TOLERANCE_MIN
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0,"mae_min":float(mae)}

def compare_scale(pred, gt) -> dict:
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1,"mae":None}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0,"mae":None}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0,"mae":None}
    try:
        vp, vg = float(pred), float(gt)
        mae = abs(vp - vg)
        return {"tp":int(mae==0),"fp":int(mae!=0),"fn":int(mae!=0),"tn":0,"mae":mae}
    except (ValueError, TypeError):
        return {"tp":0,"fp":1,"fn":1,"tn":0,"mae":None}

def compare_binary(pred, gt) -> dict:
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0}
    def to_bool(v):
        if isinstance(v, bool): return v
        s = str(v).lower()
        if s in {"true","sim","yes","1"}: return True
        if s in {"false","nao","não","no","0"}: return False
        return None
    vp, vg = to_bool(pred), to_bool(gt)
    if vp is None or vg is None:
        return {"tp":0,"fp":1,"fn":1,"tn":0}
    hit = vp == vg
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0}

def compare_categorical(pred, gt) -> dict:
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0}
    hit = normalize_cat(pred) == normalize_cat(gt)
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0}

def compare_etiologia(pred, gt) -> dict:
    """Comparação com mapeamento de variantes (Cardioembólica vs Cardioembolico, etc.)."""
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0}
    hit = normalize_etiologia(pred) == normalize_etiologia(gt)
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0}

def compare_tratamento(pred, gt) -> dict:
    """
    Comparação fuzzy para tratamento (texto livre).
    Extrai chave semântica de ambos e compara.
    Bridging: compara também cidade de origem.
    """
    pn, gn = is_null(pred), is_null(gt)
    if gn and pn:     return {"tp":0,"fp":0,"fn":0,"tn":1}
    if gn and not pn: return {"tp":0,"fp":1,"fn":0,"tn":0}
    if not gn and pn: return {"tp":0,"fp":0,"fn":1,"tn":0}
    kp, kg = _tratamento_key(pred), _tratamento_key(gt)
    # Para bridging: aceita se cidade bate; se cidade é "?" aceita parcialmente
    if kg.startswith("bridging:") and kp.startswith("bridging:"):
        cidade_gt = kg.split(":")[1]
        cidade_p  = kp.split(":")[1]
        hit = cidade_gt == cidade_p or cidade_p == "?"
    else:
        hit = kp == kg
    return {"tp":int(hit),"fp":int(not hit),"fn":int(not hit),"tn":0}

COMPARATORS = {
    **{v: compare_timestamp   for v in TIMESTAMP_VARS},
    **{v: compare_metric      for v in METRIC_VARS},
    **{v: compare_scale       for v in SCALE_VARS + NUMERIC_VARS},
    **{v: compare_binary      for v in BINARY_VARS},
    **{v: compare_categorical for v in CATEGORICAL_VARS},
    # Comparadores especializados sobrescrevem o genérico
    "etiologia_toast": compare_etiologia,
    "tratamento":      compare_tratamento,
}


# ══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO
# ══════════════════════════════════════════════════════════════════════════════

def load_ground_truth(case_dir: Path) -> dict | None:
    gt_files = list(case_dir.glob("*ground_truth*.json"))
    if not gt_files:
        print(f"  ⚠️  Sem ground truth em: {case_dir.name}")
        return None
    raw = json.loads(gt_files[0].read_text(encoding="utf-8"))
    return flatten_ground_truth(raw)

def load_cached_output(case_dir: Path) -> dict | None:
    cache = OUTPUTS_DIR / f"{case_dir.name}_output.json"
    if cache.exists():
        raw = json.loads(cache.read_text(encoding="utf-8"))
        return flatten_extractor_output(raw)
    return None

def save_cached_output(case_dir: Path, raw: dict):
    OUTPUTS_DIR.mkdir(exist_ok=True)
    cache = OUTPUTS_DIR / f"{case_dir.name}_output.json"
    cache.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# AGENTES LLM
# ══════════════════════════════════════════════════════════════════════════════

def run_agent_on_case(case_dir: Path, backend: str = "groq") -> dict:
    if str(STREAMLIT_DIR) not in sys.path:
        sys.path.insert(0, str(STREAMLIT_DIR))
    os.environ["LLM_BACKEND"] = backend
    original_dir = os.getcwd()
    os.chdir(STREAMLIT_DIR)
    try:
        txt_files   = list(case_dir.glob("*.txt"))
        carta       = next((f for f in txt_files
                            if "consulta"    not in f.name
                            and "mortalidade" not in f.name), None)
        consulta    = next((f for f in txt_files if "consulta"    in f.name), None)
        mortalidade = next((f for f in txt_files if "mortalidade" in f.name), None)

        if not carta:
            return {}

        # Agente 1: Timestamps
        from agents.extractor import extract_timestamps
        raw = extract_timestamps(carta)

        # Agente 2: Métricas
        sys.path.insert(0, str(STREAMLIT_DIR / "agents"))
        from metrics import calculate_metrics
        raw["metrics"] = calculate_metrics(raw.get("timestamps", {}))

        # Agente 3: Escalas
        raw["scales"] = {}
        try:
            from agents.scales import extract_scales
            raw["scales"]["carta"] = extract_scales(carta)
            if consulta:
                raw["scales"]["consulta"] = extract_scales(consulta)
        except Exception as e:
            print(f"    ⚠️  Escalas falharam ({case_dir.name}): {e}")

        # Agente 4: Categóricas + Mortalidade (RF6)
        raw["categorical"] = {}
        raw["mortality"]   = {"vivo_30_dias": None, "dias_obito": None, "causa_obito": None}
        try:
            from agents.categorical import extract_categorical, extract_mortality
            raw["categorical"] = extract_categorical(carta)
            if mortalidade:
                raw["mortality"] = extract_mortality(mortalidade)
        except Exception as e:
            print(f"    ⚠️  Categóricas falharam ({case_dir.name}): {e}")

        # Guarda cache só depois de todos os agentes
        save_cached_output(case_dir, raw)

        return flatten_extractor_output(raw)
    finally:
        os.chdir(original_dir)


# ══════════════════════════════════════════════════════════════════════════════
# AVALIAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def evaluate_case(predicted: dict, ground_truth: dict) -> dict:
    results = {}
    for var in ALL_VARS:
        comparator = COMPARATORS.get(var, compare_categorical)
        results[var] = comparator(predicted.get(var), ground_truth.get(var))
    return results

def aggregate_metrics(case_results: list[dict]) -> pd.DataFrame:
    agg = defaultdict(lambda: {"tp":0,"fp":0,"fn":0,"tn":0,"mae_vals":[]})
    for cr in case_results:
        for var, res in cr.items():
            agg[var]["tp"] += res.get("tp", 0)
            agg[var]["fp"] += res.get("fp", 0)
            agg[var]["fn"] += res.get("fn", 0)
            agg[var]["tn"] += res.get("tn", 0)
            for key in ("mae_min","mae"):
                v = res.get(key)
                if v is not None: agg[var]["mae_vals"].append(v)
    rows = []
    for var, a in agg.items():
        tp, fp, fn = a["tp"], a["fp"], a["fn"]
        prec   = tp/(tp+fp) if (tp+fp)>0 else None
        recall = tp/(tp+fn) if (tp+fn)>0 else None
        f1     = (2*prec*recall/(prec+recall)
                  if (prec is not None and recall is not None and (prec+recall)>0) else None)
        mae    = sum(a["mae_vals"])/len(a["mae_vals"]) if a["mae_vals"] else None
        group  = ("timestamp"   if var in TIMESTAMP_VARS   else
                  "metric"      if var in METRIC_VARS       else
                  "scale"       if var in SCALE_VARS        else
                  "binary"      if var in BINARY_VARS       else
                  "numeric"     if var in NUMERIC_VARS      else
                  "categorical")
        rows.append({
            "variable":   var, "group": group,
            "TP": tp, "FP": fp, "FN": fn, "TN": a["tn"],
            "precision":  round(prec,4)   if prec   is not None else None,
            "recall":     round(recall,4) if recall is not None else None,
            "F1":         round(f1,4)     if f1     is not None else None,
            "MAE":        round(mae,2)    if mae    is not None else None,
            "n_compared": tp+fp+fn+a["tn"],
        })
    df = pd.DataFrame(rows).set_index("variable")
    order = {"timestamp":0,"metric":1,"scale":2,"binary":3,"numeric":4,"categorical":5}
    df["_ord"] = df["group"].map(order)
    return df.sort_values(["_ord","variable"]).drop(columns="_ord")


# ══════════════════════════════════════════════════════════════════════════════
# RELATÓRIO
# ══════════════════════════════════════════════════════════════════════════════

def print_report(df: pd.DataFrame, n_cases: int, elapsed: float):
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"  RELATÓRIO DE VALIDAÇÃO — {n_cases} casos  |  {elapsed:.1f}s")
    print(sep)
    for group in ["timestamp","metric","scale","binary","numeric","categorical"]:
        sub = df[df["group"]==group]
        if sub.empty: continue
        print(f"\n▶ {group.upper()}")
        print(f"  {'Variável':<25} {'Prec':>6} {'Rec':>6} {'F1':>6} {'MAE':>8} {'TP':>5} {'FP':>5} {'FN':>5}")
        print("  " + "-"*68)
        for var, row in sub.iterrows():
            p_s   = f"{row.precision:>6.3f}" if row.precision is not None else "   n/a"
            r_s   = f"{row.recall:>6.3f}"    if row.recall    is not None else "   n/a"
            f1_s  = f"{row.F1:>6.3f}"        if row.F1        is not None else "   n/a"
            mae_s = f"{row.MAE:>8.2f}"       if row.MAE       is not None else "     n/a"
            print(f"  {var:<25} {p_s} {r_s} {f1_s} {mae_s} {row.TP:>5} {row.FP:>5} {row.FN:>5}")
    print(f"\n{sep}\n")

def save_report(df: pd.DataFrame, n_cases: int, backend: str):
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = REPORT_DIR / f"validation_{backend}_{n_cases}casos_{ts}"
    df.to_csv(f"{stem}.csv")
    try:
        with pd.ExcelWriter(f"{stem}.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Métricas")
            from openpyxl.styles import PatternFill
            ws     = writer.sheets["Métricas"]
            green  = PatternFill("solid", fgColor="C6EFCE")
            yellow = PatternFill("solid", fgColor="FFEB9C")
            red    = PatternFill("solid", fgColor="FFC7CE")
            f1_col = list(df.columns).index("F1") + 2
            for row_idx in range(2, len(df)+2):
                cell = ws.cell(row=row_idx, column=f1_col)
                val  = cell.value
                if not isinstance(val, (int, float)): continue
                cell.fill = green if val>=0.9 else (yellow if val>=0.7 else red)
    except Exception as e:
        print(f"  ⚠️  Excel não gerado ({e}) — CSV disponível.")
    print(f"\n📄 Relatório guardado em:\n   {stem}.csv\n   {stem}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--backend",   default="groq", choices=["groq","ollama"])
    parser.add_argument("--use-cache", action="store_true")
    parser.add_argument("--cases",     type=int, default=0,
                        help="Corre os primeiros N casos")
    parser.add_argument("--case",      type=str, default=None,
                        help="Corre um caso específico pelo nome (ex: caso_051_bridging)")
    args = parser.parse_args()

    if not DATA_DIR.exists():
        print(f"❌ Pasta de dados não encontrada:\n   {DATA_DIR}")
        sys.exit(1)

    all_dirs = sorted(d for d in DATA_DIR.iterdir() if d.is_dir())

    if args.case:
        case_dirs = [d for d in all_dirs if args.case in d.name]
        if not case_dirs:
            print(f"❌ Caso não encontrado: {args.case}")
            sys.exit(1)
    else:
        case_dirs = all_dirs
        if args.cases:
            case_dirs = case_dirs[:args.cases]

    print(f"\n🔍 A validar {len(case_dirs)} casos | backend={args.backend}")
    print(f"   Dados: {DATA_DIR}\n")
    t0 = time.time()

    case_results, errors = [], []

    for i, case_dir in enumerate(case_dirs, 1):
        gt = load_ground_truth(case_dir)
        if gt is None:
            errors.append({"case": case_dir.name, "error": "missing_gt"})
            continue

        pred = load_cached_output(case_dir) if args.use_cache else None
        if pred is None:
            try:
                pred = run_agent_on_case(case_dir, backend=args.backend)
            except Exception as e:
                print(f"  ❌ [{i:3d}] {case_dir.name}: {e}")
                errors.append({"case": case_dir.name, "error": str(e)})
                continue

        case_results.append(evaluate_case(pred, gt))
        print(f"  ✅ [{i:3d}/{len(case_dirs)}] {case_dir.name}")

    elapsed = time.time() - t0

    if not case_results:
        print("❌ Nenhum resultado para agregar.")
        sys.exit(1)

    df = aggregate_metrics(case_results)
    print_report(df, len(case_results), elapsed)
    save_report(df, len(case_results), args.backend)

    if errors:
        err_path = REPORT_DIR / "validation_errors.json"
        err_path.write_text(json.dumps(errors, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"⚠️  {len(errors)} erros em:\n   {err_path}")

if __name__ == "__main__":
    main()